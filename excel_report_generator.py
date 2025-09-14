"""
excel_report_generator.py

仮想通貨バックテスト結果を基に、収益性評価のための
自動Excelレポートを生成するモジュール（構想最上位レベル準拠）
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
from pandas.api.types import is_numeric_dtype
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as ExcelImage
from tempfile import NamedTemporaryFile


class ExcelReportGenerator:
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def add_summary_sheet(self):
        sheet = self.wb.create_sheet("サマリー")
        df = self.df.copy()
        df["result"] = np.where(df["profit"] > 0, "win", "lose")

        grouped = df.groupby(["symbol", "position"])
        summary = grouped.agg(
            win_rate=("result", lambda x: float(np.mean(x == "win"))),  # float()でスカラー値に変換
            pf=("profit", lambda x: float(x[x > 0].sum() / abs(x[x < 0].sum()) if any(x < 0) and x[x > 0].sum() > 0 else 0)),  # ゼロ除算対策とfloat変換
            trade_count=("profit", "count"),
            avg_profit=("profit", "mean"),
            max_dd=("profit", lambda x: float((x.cumsum() - x.cumsum().expanding().max()).min())),  # max_ddの計算を修正
        ).reset_index()

        for row in dataframe_to_rows(summary, index=False, header=True):
            sheet.append(row)

    def add_indicator_tables_like_scores(self):
        """
        Score_vs_Profit と全く同じ作りで、生指標の値幅ビンごとの
        勝率・平均利益・件数などをまとめた表を出力する。
        シート名: Indicator_vs_Profit
        """
        sheet = self.wb.create_sheet("Indicator_vs_Profit")
        df = self.df.copy()

        # 対象となる“生指標”候補（存在する列だけ使う）
        indicator_cols = [
            "RSI", "CCI", "MFI", "ADX", "ATR",
            "BB_width", "MACD_histogram", "MACD", "MACD_signal",
            "EMA_short", "EMA_long", "MA25", "plus_di14", "minus_di14"
        ]
        indicator_cols = [c for c in indicator_cols if c in df.columns]

        if not indicator_cols:
            sheet.append(["対象となる指標列が見つかりません"])
            return

        summary_data = []

        # Score_vs_Profit と同様：銘柄ごと
        for symbol in df["symbol"].dropna().unique():
            df_symbol = df[df["symbol"] == symbol]

            # Score_vs_Profit の実装はスコア名でロング/ショートを切り分けていましたが、
            # 生指標版は“ポジション別の成績も見たい”前提で long/short/all の3通りを出します。
            for position_label, position_filter in [
                ("long",  df_symbol["type"] == "long"),
                ("short", df_symbol["type"] == "short"),
                ("all",   pd.Series([True] * len(df_symbol), index=df_symbol.index))
            ]:
                filtered_df = df_symbol[position_filter]

                # 各生指標ごとに qcut=30（重複時は縮退）でビンを作成し、同じ集計形にする
                for col in indicator_cols:
                    if col not in filtered_df.columns:
                        continue

                    sub_df = filtered_df[[col, "profit"]].dropna()
                    if len(sub_df) < 10:
                        continue

                    try:
                        bins = pd.qcut(sub_df[col], q=30, duplicates='drop')
                    except ValueError:
                        # ユニーク値が足りない場合は Score_vs_Profit と同じ方針で縮退
                        unique_values = sub_df[col].nunique()
                        if unique_values < 20:
                            try:
                                bins = pd.qcut(sub_df[col], q=min(unique_values, 10), duplicates='drop')
                            except ValueError:
                                continue
                        else:
                            continue

                    grouped = sub_df.groupby(bins).agg(
                        avg_indicator=(col, "mean"),
                        avg_profit=("profit", "mean"),
                        win_rate=("profit", lambda x: np.mean(x > 0)),
                        count=("profit", "count")
                    ).reset_index()

                    # Score_vs_Profit と同じようにメタ列を追加
                    grouped.insert(0, "symbol", symbol)
                    grouped.insert(1, "indicator_feature", col)  # Score_vs_Profit の score_feature 相当
                    grouped.insert(2, "position_type", position_label)

                    summary_data.append(grouped)

        if not summary_data:
            sheet.append(["十分なデータが無く、表を作成できませんでした"])
            return

        final_df = pd.concat(summary_data, ignore_index=True)

        # Interval を文字列へ（Score_vs_Profit と同様に str() に任せる）
        def safe_convert(val):
            if isinstance(val, pd.Interval):
                return str(val)  # "(0.0906, 0.183]" のような表記
            return val

        # 出力時の列入れ替えも Score_vs_Profit と完全一致させる：
        # （clean_row[:3] + clean_row[4:8] + [clean_row[3]] + clean_row[8:]）
        # 元の列順: ["symbol","indicator_feature","position_type","bin","avg_indicator","avg_profit","win_rate","count", ...]
        # 出力後:   symbol / indicator_feature / position_type / avg_indicator / avg_profit / win_rate / count / bin / ...
        for i, row in enumerate(dataframe_to_rows(final_df, index=False, header=True)):
            clean_row = [safe_convert(cell) for cell in row]
            if len(clean_row) >= 8:
                rearranged = clean_row[:3] + clean_row[4:8] + [clean_row[3]] + clean_row[8:]
                sheet.append(rearranged)
            else:
                sheet.append(clean_row)

        # 条件付き書式の追加（★Score_vs_Profit と全く同じ列指定★）
        # コメント上は「勝率→E列 / 平均利益→F列」と書いてある実装に合わせます
        from openpyxl.formatting.rule import ColorScaleRule
        rule_win = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFAAAA',
                                mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                                end_type='percentile', end_value=100, end_color='AAFFAA')
        rule_profit = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFAAAA',
                                    mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                                    end_type='percentile', end_value=100, end_color='AAFFAA')

        # E列（5列目）とF列（6列目）に適用（Score_vs_Profit の実装と同じ指定）
        sheet.conditional_formatting.add(f'E2:E{sheet.max_row}', rule_win)     # 勝率
        sheet.conditional_formatting.add(f'F2:F{sheet.max_row}', rule_profit)  # 平均利益

    def add_charts_sheet(self):
        sheet = self.wb.create_sheet("Charts")
        df = self.df.copy()

        def plot_and_insert(x_col, y_col, anchor):
            data = df[[x_col, y_col]].dropna()
            if data.empty:
                return
            with NamedTemporaryFile(delete=False, suffix=".png") as tmpfile:
                plt.figure(figsize=(6, 4))
                sns.scatterplot(x=x_col, y=y_col, data=data, alpha=0.5)
                sns.kdeplot(data=data, x=x_col, y=y_col, fill=True, alpha=0.3, thresh=0.1)
                plt.title(f"{x_col} vs {y_col}")
                plt.tight_layout()
                plt.savefig(tmpfile.name)
                plt.close()
                img = ExcelImage(tmpfile.name)
                sheet.add_image(img, anchor)

        anchors = ["B2", "B22", "B42", "B62", "B82", "B102", "B122"]
        all_score_cols = [col for col in df.columns if col.endswith("_long") or col.endswith("_short")]
        plotted = 0

        for col in all_score_cols:
            if col in df.columns and plotted < len(anchors):
                plot_and_insert(col, "profit", anchors[plotted])
                plotted += 1

        # 表としても出力
        summary_sheet = self.wb.create_sheet("Score_vs_Profit")
        summary_data = []
        for symbol in df["symbol"].dropna().unique():
            df_symbol = df[df["symbol"] == symbol]
            for col in all_score_cols:
                if col in df_symbol.columns:
                    # ★修正: スコア名からポジション種別を判定★
                    if col.endswith('_long'):
                        # ロング用スコアの場合、ロングポジションのみを対象
                        position_filter = df_symbol['type'] == 'long'
                        position_label = 'long'
                    elif col.endswith('_short'):
                        # ショート用スコアの場合、ショートポジションのみを対象
                        position_filter = df_symbol['type'] == 'short'
                        position_label = 'short'
                    else:
                        # その他のスコア（buy_score, sell_scoreなど）は全ポジション対象
                        position_filter = pd.Series([True] * len(df_symbol))
                        position_label = 'all'
                    
                    # フィルタリングされたデータを取得
                    filtered_df = df_symbol[position_filter]
                    sub_df = filtered_df[[col, "profit"]].dropna()
                    
                    if len(sub_df) < 10:  # 最低限のデータ数をチェック
                        continue
                    
                    # ★修正: 10分割から20分割に変更★
                    try:
                        bins = pd.qcut(sub_df[col], q=30, duplicates='drop')
                    except ValueError:
                        # 20分割でユニークな値が足りない場合は、データの長さに応じて調整
                        unique_values = sub_df[col].nunique()
                        if unique_values < 20:
                            # ユニークな値が20未満の場合は、可能な最大分割数を使用
                            try:
                                bins = pd.qcut(sub_df[col], q=min(unique_values, 10), duplicates='drop')
                            except ValueError:
                                continue
                        else:
                            continue
                    
                    grouped = sub_df.groupby(bins).agg(
                        avg_score=(col, "mean"),
                        avg_profit=("profit", "mean"),
                        win_rate=("profit", lambda x: np.mean(x > 0)),
                        count=("profit", "count")
                    ).reset_index()
                    grouped.insert(0, "symbol", symbol)
                    grouped.insert(1, "score_feature", col)
                    grouped.insert(2, "position_type", position_label)  # ★追加: ポジション種別を明示★
                    summary_data.append(grouped)

        if summary_data:
            final_df = pd.concat(summary_data, ignore_index=True)

            def safe_convert(val):
                if isinstance(val, pd.Interval):
                    return str(val)
                return val

            for i, row in enumerate(dataframe_to_rows(final_df, index=False, header=True)):
                clean_row = [safe_convert(cell) for cell in row]
                if len(clean_row) >= 8:  # ★修正: position_type列が追加されたので8列以上に変更★
                    rearranged = clean_row[:3] + clean_row[4:8] + [clean_row[3]] + clean_row[8:]
                    summary_sheet.append(rearranged)
                else:
                    summary_sheet.append(clean_row)

            # 条件付き書式の追加（勝率・平均利益列）
            from openpyxl.formatting.rule import ColorScaleRule

            # 勝率 → E列（position_type列が追加されたため列番号が変更）
            rule_win = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFAAAA',
                                      mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                                      end_type='percentile', end_value=100, end_color='AAFFAA')
            summary_sheet.conditional_formatting.add(f'E2:E{summary_sheet.max_row}', rule_win)

            # 平均利益 → F列（position_type列が追加されたため列番号が変更）
            rule_profit = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFAAAA',
                                         mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                                         end_type='percentile', end_value=100, end_color='AAFFAA')
            summary_sheet.conditional_formatting.add(f'F2:F{summary_sheet.max_row}', rule_profit)


    def add_score_analysis_sheet(self):
        sheet = self.wb.create_sheet("スコア分析")
        df = self.df.copy()

        score_cols = []
        if "buy_score" in df.columns:
            score_cols.append("buy_score")
        if "sell_score" in df.columns:
            score_cols.append("sell_score")

        if not score_cols:
            sheet.append(["スコア列が見つかりません"])
            return

        try:
            for symbol in df["symbol"].dropna().unique():
                df_symbol = df[df["symbol"] == symbol]
                for score_col in score_cols:
                    # ★修正: buy_scoreはlongポジション、sell_scoreはshortポジション用★
                    if score_col == "buy_score":
                        # buy_scoreの場合、ロングポジションのみを対象
                        position_filter = df_symbol['type'] == 'long'
                        position_label = 'long'
                    elif score_col == "sell_score":
                        # sell_scoreの場合、ショートポジションのみを対象
                        position_filter = df_symbol['type'] == 'short'
                        position_label = 'short'
                    else:
                        # その他のスコアは全ポジション対象
                        position_filter = pd.Series([True] * len(df_symbol))
                        position_label = 'all'
                    
                    # フィルタリングされたデータを取得
                    filtered_df = df_symbol[position_filter]
                    valid_data = filtered_df[filtered_df[score_col].notna()].copy()
                    
                    if len(valid_data) < 10:  # 最低限のデータ数をチェック
                        continue

                    # ★修正: スコア分析シートも20分割に変更★
                    def get_score_bin(score):
                        if pd.isna(score):
                            return "N/A"
                        elif score < 0.05:
                            return "0.00-0.05"
                        elif score < 0.10:
                            return "0.05-0.10"
                        elif score < 0.15:
                            return "0.10-0.15"
                        elif score < 0.20:
                            return "0.15-0.20"
                        elif score < 0.25:
                            return "0.20-0.25"
                        elif score < 0.30:
                            return "0.25-0.30"
                        elif score < 0.35:
                            return "0.30-0.35"
                        elif score < 0.40:
                            return "0.35-0.40"
                        elif score < 0.45:
                            return "0.40-0.45"
                        elif score < 0.50:
                            return "0.45-0.50"
                        elif score < 0.55:
                            return "0.50-0.55"
                        elif score < 0.60:
                            return "0.55-0.60"
                        elif score < 0.65:
                            return "0.60-0.65"
                        elif score < 0.70:
                            return "0.65-0.70"
                        elif score < 0.75:
                            return "0.70-0.75"
                        elif score < 0.80:
                            return "0.75-0.80"
                        elif score < 0.85:
                            return "0.80-0.85"
                        elif score < 0.90:
                            return "0.85-0.90"
                        elif score < 0.95:
                            return "0.90-0.95"
                        else:
                            return "0.95-1.00"

                    valid_data["score_bin"] = valid_data[score_col].apply(get_score_bin)

                    analysis = valid_data.groupby("score_bin").agg(
                        勝率=("profit", lambda x: float(np.mean(x > 0))),
                        平均利益=("profit", "mean"),
                        取引回数=("profit", "count"),
                        平均スコア=(score_col, "mean")
                    ).reset_index()
                    analysis.insert(0, "スコア種別", score_col)
                    analysis.insert(0, "通貨ペア", symbol)
                    analysis.insert(1, "ポジション種別", position_label)  # ★追加: ポジション種別を明示★

                    if sheet.max_row > 1:
                        sheet.append([])  # 空行で区切る

                    for row in dataframe_to_rows(analysis, index=False, header=True):
                        sheet.append(row)

        except Exception as e:
            sheet.append([f"スコア分析でエラー: {str(e)}"])


    def add_market_condition_sheet(self):
        sheet = self.wb.create_sheet("市場状況")
        df = self.df.copy()

        required_cols = ["ADX", "ATR", "RSI", "symbol"]
        if all(col in df.columns for col in required_cols):
            try:
                df = df.dropna(subset=required_cols + ["profit"]).copy()

                # 各指標を5段階にビニング
                adx_labels = ["very_low", "low", "medium", "high", "very_high"]
                atr_labels = ["very_low", "low", "medium", "high", "very_high"]
                rsi_labels = ["very_low", "low", "medium", "high", "very_high"]

                df["ADX_bin"] = pd.qcut(df["ADX"], q=5, labels=adx_labels)
                df["ATR_bin"] = pd.qcut(df["ATR"], q=5, labels=atr_labels)
                df["RSI_bin"] = pd.qcut(df["RSI"], q=5, labels=rsi_labels)

                # グルーピング
                grouped = df.groupby(["symbol", "ADX_bin", "ATR_bin", "RSI_bin"]).agg(
                    win_rate=("profit", lambda x: float(np.mean(x > 0))),
                    avg_profit=("profit", "mean"),
                    count=("profit", "count")
                ).reset_index()

                for row in dataframe_to_rows(grouped, index=False, header=True):
                    sheet.append(row)

            except Exception as e:
                sheet.append([f"市場状況分析でエラー: {str(e)}"])
        else:
            sheet.append(["ADX, ATR, RSI, または symbol が見つかりません"])


    def add_symbol_sheets(self):
        for symbol, sdf in self.df.groupby("symbol"):
            sheet = self.wb.create_sheet(f"{symbol}")
            sdf = sdf.copy()
            sdf["result"] = np.where(sdf["profit"] > 0, "win", "lose")

            summary = sdf.groupby("position").agg(
                win_rate=("result", lambda x: np.mean(x == "win")),
                avg_profit=("profit", "mean"),
                count=("profit", "count"),
                avg_duration=("holding_period", "mean") if "holding_period" in sdf.columns else ("profit", "count")
            ).reset_index()

            sheet.append(["ポジション別の成績"])
            for row in dataframe_to_rows(summary, index=False, header=True):
                sheet.append(row)

            # 勝ち/負け特徴比較（拡張版）
            sheet.append([])
            sheet.append(["勝ち/負けトレードの特徴比較"])
            features = ["buy_score", "sell_score", "ATR", "ADX"]
            if "holding_period" in sdf.columns:
                features.append("holding_period")

            available_features = [col for col in features if col in sdf.columns]
            if available_features:
                compare = sdf.groupby("result")[available_features].mean().reset_index()
                for row in dataframe_to_rows(compare, index=False, header=True):
                    sheet.append(row)

            sheet.append([])
            sheet.append(["トレード一覧"])
            for row in dataframe_to_rows(sdf.head(50), index=False, header=True):
                sheet.append(row)


    def save(self, path: str):
        self.wb.save(path)